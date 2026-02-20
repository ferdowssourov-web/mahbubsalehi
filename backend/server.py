from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import jwt
import hashlib
import shutil
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'salehi-admin-secret-2025')
JWT_ALGORITHM = "HS256"

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()


# ======== Models ========

class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class ContactMessage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    email: Optional[str] = ""
    message: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ContactMessageCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    message: str

class Activity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    category: str
    image_url: str = ""
    content: str = ""
    date: str = ""
    is_published: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ActivityCreate(BaseModel):
    title: str
    category: str
    image_url: str = ""
    content: str = ""
    date: str = ""
    is_published: bool = True

class ActivityUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    image_url: Optional[str] = None
    content: Optional[str] = None
    date: Optional[str] = None
    is_published: Optional[bool] = None

class AdminLogin(BaseModel):
    username: str
    password: str

class AdminLoginResponse(BaseModel):
    token: str
    username: str

# Public Opinion (জনতার মতামত) Models
class PublicOpinion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    area: str = ""
    opinion: str
    rating: int = 5
    is_approved: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PublicOpinionCreate(BaseModel):
    name: str
    phone: str
    area: str = ""
    opinion: str
    rating: int = 5

class PublicOpinionUpdate(BaseModel):
    is_approved: Optional[bool] = None

# Meeting Registration Models
class MeetingRegistration(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    area: str = ""
    subject: str = ""
    message: str = ""
    status: str = "pending"  # pending, approved, completed, cancelled
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MeetingRegistrationCreate(BaseModel):
    name: str
    phone: str
    area: str = ""
    subject: str = ""
    message: str = ""

class MeetingRegistrationUpdate(BaseModel):
    status: Optional[str] = None

# Countdown Settings Model
class CountdownSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "countdown_settings"
    is_active: bool = False
    target_date: str = ""  # ISO format date string
    title: str = "তারুণ্যের মুখোমুখির পরবর্তী সময়"

class CountdownSettingsUpdate(BaseModel):
    is_active: Optional[bool] = None
    target_date: Optional[str] = None
    title: Optional[str] = None

# Meeting Address Settings Model
class MeetingAddressSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "meeting_address_settings"
    address: str = "উলিপুর, কুড়িগ্রাম, বাংলাদেশ"
    description: str = "ব্যারিস্টার মাহবুবুল আলম সালেহী উলিপুরের প্রতিটি মানুষের কথা শুনতে চান। আপনার সমস্যা, পরামর্শ বা মতামত জানাতে নিচের ফর্মটি পূরণ করে রেজিষ্ট্রেশন সম্পূর্ণ করুন।"

class MeetingAddressSettingsUpdate(BaseModel):
    address: Optional[str] = None
    description: Optional[str] = None

# Gallery Image Models
class GalleryImage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    image_url: str
    caption: str = ""
    category: str  # "local" or "international"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class GalleryImageCreate(BaseModel):
    image_url: str
    caption: str = ""
    category: str  # "local" or "international"


# ======== Auth Helpers ========

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def create_token(username: str) -> str:
    payload = {
        "sub": username,
        "iat": datetime.now(timezone.utc).timestamp(),
        "exp": (datetime.now(timezone.utc).timestamp()) + 86400  # 24 hours
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ======== Seed Default Admin ========

async def seed_admin():
    existing = await db.admins.find_one({"username": "admin"})
    if not existing:
        await db.admins.insert_one({
            "username": "admin",
            "password_hash": hash_password("admin123"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logging.info("Default admin created: admin / admin123")

async def seed_activities():
    count = await db.activities.count_documents({})
    if count == 0:
        seed_data = [
            {
                "id": str(uuid.uuid4()),
                "title": "ব্যারিস্টার মাহবুবুল আলম সালেহী: উলিপুরের এক অদম্য সংগ্রামীর গল্প",
                "category": "ব্যক্তিত্ব",
                "image_url": "https://media.bdji.org/images/1211.original.format-webp.webp",
                "content": "ব্যারিস্টার মাহবুবুল আলম সালেহী উলিপুরের এক অদম্য সংগ্রামী যিনি আন্তর্জাতিক মঞ্চে বাংলাদেশের গণতন্ত্র ও মানবাধিকারের পক্ষে কাজ করেছেন।",
                "date": "২০২৫",
                "is_published": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "title": "জামায়াতে ইসলামীর কেন্দ্রীয় নির্বাহী পরিষদের বৈঠক",
                "category": "রাজনীতি",
                "image_url": "https://media.bdji.org/images/557585451_122151827564749601_11561.original.format-webp.webp",
                "content": "জামায়াতে ইসলামীর কেন্দ্রীয় নির্বাহী পরিষদের বৈঠকে বিভিন্ন গুরুত্বপূর্ণ বিষয় নিয়ে আলোচনা হয়।",
                "date": "২০২৫",
                "is_published": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "title": "কুড়িগ্রামে জামায়াতের মিছিল ও জেলা প্রশাসকের কাছে স্মারকলিপি প্রদান",
                "category": "কার্যক্রম",
                "image_url": "https://media.bdji.org/images/4242.original.format-webp.webp",
                "content": "কুড়িগ্রামে জামায়াতের মিছিল ও জেলা প্রশাসকের কাছে স্মারকলিপি প্রদান করা হয়েছে।",
                "date": "২০২৫",
                "is_published": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "title": "নিন্দা ও প্রতিবাদ বার্তা",
                "category": "বিবৃতি",
                "image_url": "https://media.bdji.org/images/121212.original.format-webp.webp",
                "content": "সাম্প্রতিক ঘটনায় নিন্দা ও প্রতিবাদ জানিয়ে বিবৃতি প্রদান করেছেন ব্যারিস্টার সালেহী।",
                "date": "২০২৫",
                "is_published": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "title": "সমৃদ্ধ উলিপুর গড়ার লক্ষ্যে রংপুরে উলিপুরবাসীর মতবিনিময় সভা অনুষ্ঠিত",
                "category": "সভা",
                "image_url": "https://media.bdji.org/images/557602506_122136658088904504_46541.original.format-webp.webp",
                "content": "সমৃদ্ধ উলিপুর গড়ার লক্ষ্যে রংপুরে উলিপুরবাসীর মতবিনিময় সভা সফলভাবে অনুষ্ঠিত হয়েছে।",
                "date": "২০২৫",
                "is_published": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.activities.insert_many(seed_data)
        logging.info(f"Seeded {len(seed_data)} activities")


# ======== Routes ========

@api_router.get("/")
async def root():
    return {"message": "Hello World"}

# Status endpoints
@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks

# Contact endpoints
@api_router.post("/contact", response_model=ContactMessage)
async def create_contact_message(input: ContactMessageCreate):
    contact_dict = input.model_dump()
    contact_obj = ContactMessage(**contact_dict)
    doc = contact_obj.model_dump()
    _ = await db.contact_messages.insert_one(doc)
    return contact_obj

@api_router.get("/contact", response_model=List[ContactMessage])
async def get_contact_messages():
    messages = await db.contact_messages.find({}, {"_id": 0}).to_list(1000)
    return messages

# ======== Admin Auth ========

@api_router.post("/admin/login", response_model=AdminLoginResponse)
async def admin_login(input: AdminLogin):
    admin = await db.admins.find_one({"username": input.username}, {"_id": 0})
    if not admin or admin["password_hash"] != hash_password(input.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(input.username)
    return AdminLoginResponse(token=token, username=input.username)

@api_router.get("/admin/me")
async def admin_me(username: str = Depends(verify_token)):
    return {"username": username}

# ======== Activities CRUD ========

@api_router.get("/activities", response_model=List[Activity])
async def get_activities(published_only: bool = True):
    query = {"is_published": True} if published_only else {}
    activities = await db.activities.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return activities

@api_router.get("/activities/{activity_id}", response_model=Activity)
async def get_activity(activity_id: str):
    activity = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    return activity

@api_router.post("/activities", response_model=Activity)
async def create_activity(input: ActivityCreate, username: str = Depends(verify_token)):
    activity_dict = input.model_dump()
    activity_obj = Activity(**activity_dict)
    doc = activity_obj.model_dump()
    _ = await db.activities.insert_one(doc)
    return activity_obj

@api_router.put("/activities/{activity_id}", response_model=Activity)
async def update_activity(activity_id: str, input: ActivityUpdate, username: str = Depends(verify_token)):
    existing = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.activities.update_one({"id": activity_id}, {"$set": update_data})
    updated = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    return Activity(**updated)

@api_router.delete("/activities/{activity_id}")
async def delete_activity(activity_id: str, username: str = Depends(verify_token)):
    result = await db.activities.delete_one({"id": activity_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activity not found")
    return {"message": "Activity deleted"}

# ======== Admin: Contact Messages ========

@api_router.get("/admin/contacts", response_model=List[ContactMessage])
async def get_admin_contacts(username: str = Depends(verify_token)):
    messages = await db.contact_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return messages

@api_router.delete("/admin/contacts/{contact_id}")
async def delete_contact(contact_id: str, username: str = Depends(verify_token)):
    result = await db.contact_messages.delete_one({"id": contact_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    return {"message": "Contact deleted"}


# ======== Public Opinion (জনতার মতামত) ========

@api_router.post("/opinions", response_model=PublicOpinion)
async def create_opinion(input: PublicOpinionCreate):
    opinion_dict = input.model_dump()
    opinion_obj = PublicOpinion(**opinion_dict)
    doc = opinion_obj.model_dump()
    _ = await db.public_opinions.insert_one(doc)
    return opinion_obj

@api_router.get("/opinions", response_model=List[PublicOpinion])
async def get_opinions(approved_only: bool = True):
    query = {"is_approved": True} if approved_only else {}
    opinions = await db.public_opinions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return opinions

@api_router.get("/admin/opinions", response_model=List[PublicOpinion])
async def get_admin_opinions(username: str = Depends(verify_token)):
    opinions = await db.public_opinions.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return opinions

@api_router.put("/admin/opinions/{opinion_id}", response_model=PublicOpinion)
async def update_opinion(opinion_id: str, input: PublicOpinionUpdate, username: str = Depends(verify_token)):
    existing = await db.public_opinions.find_one({"id": opinion_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Opinion not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    await db.public_opinions.update_one({"id": opinion_id}, {"$set": update_data})
    updated = await db.public_opinions.find_one({"id": opinion_id}, {"_id": 0})
    return PublicOpinion(**updated)

@api_router.delete("/admin/opinions/{opinion_id}")
async def delete_opinion(opinion_id: str, username: str = Depends(verify_token)):
    result = await db.public_opinions.delete_one({"id": opinion_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Opinion not found")
    return {"message": "Opinion deleted"}


# ======== Gallery (গ্যালারি) ========

@api_router.get("/gallery", response_model=List[GalleryImage])
async def get_gallery_images(category: Optional[str] = None):
    query = {"category": category} if category else {}
    images = await db.gallery_images.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return images

@api_router.post("/admin/gallery", response_model=GalleryImage)
async def create_gallery_image(input: GalleryImageCreate, username: str = Depends(verify_token)):
    if input.category not in ["local", "international"]:
        raise HTTPException(status_code=400, detail="Category must be 'local' or 'international'")
    image_dict = input.model_dump()
    image_obj = GalleryImage(**image_dict)
    doc = image_obj.model_dump()
    _ = await db.gallery_images.insert_one(doc)
    return image_obj

@api_router.get("/admin/gallery", response_model=List[GalleryImage])
async def get_admin_gallery(username: str = Depends(verify_token)):
    images = await db.gallery_images.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return images

@api_router.delete("/admin/gallery/{image_id}")
async def delete_gallery_image(image_id: str, username: str = Depends(verify_token)):
    result = await db.gallery_images.delete_one({"id": image_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Image not found")
    return {"message": "Image deleted"}


# ======== Meeting Registration (সাক্ষাৎ রেজিষ্ট্রেশন) ========

@api_router.post("/registrations", response_model=MeetingRegistration)
async def create_registration(input: MeetingRegistrationCreate):
    reg_dict = input.model_dump()
    reg_obj = MeetingRegistration(**reg_dict)
    doc = reg_obj.model_dump()
    _ = await db.meeting_registrations.insert_one(doc)
    return reg_obj

@api_router.get("/admin/registrations", response_model=List[MeetingRegistration])
async def get_admin_registrations(username: str = Depends(verify_token)):
    registrations = await db.meeting_registrations.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return registrations

@api_router.put("/admin/registrations/{reg_id}", response_model=MeetingRegistration)
async def update_registration(reg_id: str, input: MeetingRegistrationUpdate, username: str = Depends(verify_token)):
    existing = await db.meeting_registrations.find_one({"id": reg_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if update_data.get("status") and update_data["status"] not in ["pending", "approved", "completed", "cancelled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    await db.meeting_registrations.update_one({"id": reg_id}, {"$set": update_data})
    updated = await db.meeting_registrations.find_one({"id": reg_id}, {"_id": 0})
    return MeetingRegistration(**updated)

@api_router.delete("/admin/registrations/{reg_id}")
async def delete_registration(reg_id: str, username: str = Depends(verify_token)):
    result = await db.meeting_registrations.delete_one({"id": reg_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registration not found")
    return {"message": "Registration deleted"}

@api_router.get("/admin/registrations/export/excel")
async def export_registrations_excel(username: str = Depends(verify_token)):
    """Export all meeting registrations as Excel file"""
    try:
        # Fetch all registrations
        registrations = await db.meeting_registrations.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "সাক্ষাৎ রেজিষ্ট্রেশন"
        
        # Define headers with Bengali labels
        headers = [
            "ক্রমিক নং",
            "নাম",
            "মোবাইল নম্বর",
            "এলাকা",
            "বিষয়",
            "বার্তা",
            "স্ট্যাটাস",
            "তারিখ ও সময়"
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="064e3b", end_color="064e3b", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Status translation
        status_map = {
            "pending": "অপেক্ষমাণ",
            "approved": "অনুমোদিত",
            "completed": "সম্পন্ন",
            "cancelled": "বাতিল"
        }
        
        # Write data rows
        for row_num, reg in enumerate(registrations, 2):
            # Format created_at timestamp
            try:
                created_dt = datetime.fromisoformat(reg.get("created_at", ""))
                formatted_date = created_dt.strftime("%d/%m/%Y %I:%M %p")
            except:
                formatted_date = reg.get("created_at", "")
            
            row_data = [
                row_num - 1,  # Serial number
                reg.get("name", ""),
                reg.get("phone", ""),
                reg.get("area", ""),
                reg.get("subject", ""),
                reg.get("message", ""),
                status_map.get(reg.get("status", "pending"), "অপেক্ষমাণ"),
                formatted_date
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Adjust column widths
        column_widths = [10, 25, 18, 20, 30, 40, 15, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename with current date
        filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        logging.error(f"Error exporting registrations: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ======== Countdown Settings ========

@api_router.get("/countdown", response_model=CountdownSettings)
async def get_countdown_settings():
    settings = await db.countdown_settings.find_one({"id": "countdown_settings"}, {"_id": 0})
    if not settings:
        default = CountdownSettings()
        await db.countdown_settings.insert_one(default.model_dump())
        return default
    return CountdownSettings(**settings)

@api_router.put("/admin/countdown", response_model=CountdownSettings)
async def update_countdown_settings(input: CountdownSettingsUpdate, username: str = Depends(verify_token)):
    existing = await db.countdown_settings.find_one({"id": "countdown_settings"}, {"_id": 0})
    if not existing:
        default = CountdownSettings()
        await db.countdown_settings.insert_one(default.model_dump())
        existing = default.model_dump()
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    await db.countdown_settings.update_one({"id": "countdown_settings"}, {"$set": update_data})
    updated = await db.countdown_settings.find_one({"id": "countdown_settings"}, {"_id": 0})
    return CountdownSettings(**updated)

# ======== Meeting Address Settings ========

@api_router.get("/meeting-address", response_model=MeetingAddressSettings)
async def get_meeting_address_settings():
    """Get meeting address and description"""
    settings = await db.meeting_address_settings.find_one({"id": "meeting_address_settings"}, {"_id": 0})
    if not settings:
        default = MeetingAddressSettings()
        await db.meeting_address_settings.insert_one(default.model_dump())
        return default
    return MeetingAddressSettings(**settings)

@api_router.put("/admin/meeting-address", response_model=MeetingAddressSettings)
async def update_meeting_address_settings(input: MeetingAddressSettingsUpdate, username: str = Depends(verify_token)):
    """Update meeting address and description (Admin only)"""
    existing = await db.meeting_address_settings.find_one({"id": "meeting_address_settings"}, {"_id": 0})
    if not existing:
        default = MeetingAddressSettings()
        await db.meeting_address_settings.insert_one(default.model_dump())
        existing = default.model_dump()
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    await db.meeting_address_settings.update_one({"id": "meeting_address_settings"}, {"$set": update_data})
    updated = await db.meeting_address_settings.find_one({"id": "meeting_address_settings"}, {"_id": 0})
    return MeetingAddressSettings(**updated)


# ======== Image Upload ========

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg'}

@api_router.post("/upload")
async def upload_image(file: UploadFile = File(...), username: str = Depends(verify_token)):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type not allowed. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")
    
    file_id = str(uuid.uuid4())
    filename = f"{file_id}{ext}"
    filepath = UPLOADS_DIR / filename
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Build the URL relative to the API
    file_url = f"/api/uploads/{filename}"
    return {"url": file_url, "filename": filename}


# Include router
app.include_router(api_router)

# Mount uploads as static files
app.mount("/api/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    await seed_admin()
    await seed_activities()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
